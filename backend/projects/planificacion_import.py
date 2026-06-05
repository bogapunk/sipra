"""
Importación masiva de planificación anual desde Excel (.xlsx) o CSV.

Jerarquía: Eje → Plan → Programa → Objetivo → Proyecto → Indicador
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date
from typing import Any

from django.db import transaction

from .models import Eje, Plan, Programa, ObjetivoEstrategico, Proyecto, ProyectoObjetivo, Indicador

LEGACY_ANIO = 2026

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    'id_eje': ('id_eje', 'id eje', 'codigo_eje', 'codigo eje', 'eje_id', 'eje id'),
    'eje': ('eje', 'nombre_eje', 'nombre eje'),
    'id_plan': ('id_plan', 'id plan', 'codigo_plan', 'codigo plan', 'plan_id', 'plan id'),
    'plan': ('plan', 'nombre_plan', 'nombre plan'),
    'proposito_plan': (
        'proposito_plan', 'proposito plan', 'proposito_politica_publica',
        'proposito politica publica', 'proposito',
    ),
    'vision_plan': ('vision_plan', 'vision plan', 'vision_estrategica', 'vision estrategica', 'vision'),
    'id_programa': ('id_programa', 'id programa', 'codigo_programa', 'codigo programa', 'programa_id'),
    'programa': ('programa', 'nombre_programa', 'nombre programa'),
    'objetivo': ('objetivo', 'objetivo_estrategico', 'objetivo estrategico', 'descripcion_objetivo'),
    'proyecto': ('proyecto', 'nombre_proyecto', 'nombre proyecto'),
    'indicador': ('indicador', 'indicador_descripcion', 'descripcion_indicador', 'descripcion indicador'),
    'unidad_medida': ('unidad_medida', 'unidad medida', 'unidad'),
    'frecuencia': ('frecuencia',),
}


def _normalize_header(value: Any) -> str:
    text = str(value or '').strip().lower()
    text = re.sub(r'\s+', ' ', text)
    return text


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    alias_lookup: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_lookup[_normalize_header(alias)] = canonical
    for idx, header in enumerate(raw_headers):
        key = alias_lookup.get(_normalize_header(header))
        if key:
            mapped[idx] = key
    return mapped


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_int(value: Any) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def eje_pk(anio: int, codigo: int) -> int:
    if anio == LEGACY_ANIO and codigo <= 6:
        return codigo
    return anio * 100 + codigo


def plan_pk(anio: int, codigo: int) -> int:
    if anio == LEGACY_ANIO and codigo <= 9:
        return codigo
    return anio * 100 + codigo


def programa_pk(anio: int, local: str) -> str:
    local = str(local or '').strip()
    if not local:
        return local
    if anio == LEGACY_ANIO and not local.startswith(str(anio)):
        return local
    if local.startswith(f'{anio}.'):
        return local
    return f'{anio}.{local}'


def parse_rows_from_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    header_map = _map_headers(raw_headers)
    if not header_map:
        raise ValueError(
            'No se reconocieron columnas válidas. Use la plantilla con encabezados como '
            'id_eje, eje, id_plan, plan, id_programa, programa, objetivo, proyecto, indicador.'
        )
    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        if not any(str(c or '').strip() for c in raw):
            continue
        item: dict[str, str] = {}
        for idx, key in header_map.items():
            if idx < len(raw):
                item[key] = _cell_str(raw[idx])
        rows.append(item)
    return rows


def parse_rows_from_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('openpyxl no está instalado en el servidor.') from exc
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    header_map = _map_headers(list(raw_headers))
    if not header_map:
        raise ValueError(
            'No se reconocieron columnas válidas. Use la plantilla con encabezados como '
            'id_eje, eje, id_plan, plan, id_programa, programa, objetivo, proyecto, indicador.'
        )
    rows: list[dict[str, str]] = []
    for raw in rows_iter:
        if not raw or not any(str(c or '').strip() for c in raw):
            continue
        item: dict[str, str] = {}
        for idx, key in header_map.items():
            if idx < len(raw):
                item[key] = _cell_str(raw[idx])
        rows.append(item)
    return rows


def parse_planificacion_file(filename: str, content: bytes) -> list[dict[str, str]]:
    ext = filename.rsplit('.', 1)[-1].lower() if filename else ''
    if ext == 'csv':
        return parse_rows_from_csv(content)
    if ext in ('xlsx', 'xlsm'):
        return parse_rows_from_xlsx(content)
    raise ValueError('Formato no soportado. Use archivos .xlsx o .csv')


def _fill_down(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    keys = list(COLUMN_ALIASES.keys())
    last: dict[str, str] = {}
    filled: list[dict[str, str]] = []
    for row in rows:
        merged = dict(last)
        for key in keys:
            val = row.get(key, '')
            if val:
                merged[key] = val
        if merged.get('id_eje') or merged.get('eje'):
            last = merged
            filled.append(merged)
    return filled


@transaction.atomic
def import_planificacion_rows(rows: list[dict[str, str]], anio: int, usuario) -> dict[str, int]:
    if anio < 2000 or anio > 2100:
        raise ValueError('El año debe estar entre 2000 y 2100.')

    stats = {
        'filas_procesadas': 0,
        'ejes': 0,
        'planes': 0,
        'programas': 0,
        'objetivos': 0,
        'proyectos': 0,
        'indicadores': 0,
    }
    created_flags = {
        'ejes': set(),
        'planes': set(),
        'programas': set(),
        'objetivos': set(),
        'proyectos': set(),
        'indicadores': set(),
    }

    filled_rows = _fill_down(rows)
    if not filled_rows:
        raise ValueError('El archivo no contiene filas de datos.')

    fecha_inicio = date(anio, 1, 1)
    fecha_fin = date(anio, 12, 31)

    for row in filled_rows:
        stats['filas_procesadas'] += 1

        codigo_eje = _parse_int(row.get('id_eje'))
        nombre_eje = row.get('eje', '')
        if codigo_eje is None and not nombre_eje:
            continue
        if codigo_eje is None:
            codigo_eje = 1

        pk_eje = eje_pk(anio, codigo_eje)
        eje, created = Eje.objects.update_or_create(
            id_eje=pk_eje,
            defaults={'anio': anio, 'nombre_eje': nombre_eje or f'Eje {codigo_eje}'},
        )
        if created and pk_eje not in created_flags['ejes']:
            created_flags['ejes'].add(pk_eje)
            stats['ejes'] += 1

        codigo_plan = _parse_int(row.get('id_plan'))
        nombre_plan = row.get('plan', '')
        if codigo_plan is None and not nombre_plan:
            continue
        if codigo_plan is None:
            codigo_plan = codigo_eje

        pk_plan = plan_pk(anio, codigo_plan)
        plan, created = Plan.objects.update_or_create(
            id_plan=pk_plan,
            defaults={
                'eje': eje,
                'nombre_plan': nombre_plan or f'Plan {codigo_plan}',
                'proposito_politica_publica': row.get('proposito_plan', ''),
                'vision_estrategica': row.get('vision_plan', ''),
            },
        )
        if created and pk_plan not in created_flags['planes']:
            created_flags['planes'].add(pk_plan)
            stats['planes'] += 1

        local_programa = row.get('id_programa', '')
        nombre_programa = row.get('programa', '')
        if not local_programa and not nombre_programa:
            continue
        if not local_programa:
            local_programa = f'{codigo_plan}.1'

        pk_programa = programa_pk(anio, local_programa)
        programa, created = Programa.objects.update_or_create(
            id_programa=pk_programa,
            defaults={
                'plan': plan,
                'nombre_programa': nombre_programa or f'Programa {local_programa}',
            },
        )
        if created and pk_programa not in created_flags['programas']:
            created_flags['programas'].add(pk_programa)
            stats['programas'] += 1

        desc_obj = row.get('objetivo', '').strip()
        if not desc_obj:
            continue

        objetivo, created = ObjetivoEstrategico.objects.get_or_create(
            programa=programa,
            descripcion=desc_obj,
        )
        obj_key = (pk_programa, desc_obj)
        if created and obj_key not in created_flags['objetivos']:
            created_flags['objetivos'].add(obj_key)
            stats['objetivos'] += 1

        nombre_proyecto = row.get('proyecto', '').strip()
        proyecto = None
        if nombre_proyecto:
            proyecto = Proyecto.objects.filter(
                nombre=nombre_proyecto,
                objetivo_estrategico=objetivo,
            ).first()
            if not proyecto:
                proyecto = Proyecto.objects.create(
                    nombre=nombre_proyecto,
                    descripcion='',
                    fecha_inicio=fecha_inicio,
                    fecha_fin_estimada=fecha_fin,
                    creado_por=usuario,
                    programa=programa,
                    objetivo_estrategico=objetivo,
                )
                stats['proyectos'] += 1
                created_flags['proyectos'].add(proyecto.id)
            else:
                proyecto.programa = programa
                proyecto.objetivo_estrategico = objetivo
                proyecto.save(update_fields=['programa', 'objetivo_estrategico'])

            ProyectoObjetivo.objects.get_or_create(
                proyecto=proyecto,
                objetivo=objetivo,
                defaults={'estado_avance': ProyectoObjetivo.ESTADO_NO_INICIADO},
            )

        desc_ind = row.get('indicador', '').strip()
        if desc_ind and proyecto:
            ind, created = Indicador.objects.get_or_create(
                proyecto=proyecto,
                descripcion=desc_ind,
                defaults={
                    'unidad_medida': row.get('unidad_medida', ''),
                    'frecuencia': row.get('frecuencia', ''),
                },
            )
            if not created:
                changed = False
                unidad = row.get('unidad_medida', '')
                frec = row.get('frecuencia', '')
                if unidad and ind.unidad_medida != unidad:
                    ind.unidad_medida = unidad
                    changed = True
                if frec and ind.frecuencia != frec:
                    ind.frecuencia = frec
                    changed = True
                if changed:
                    ind.save()
            else:
                ind_key = (proyecto.id, desc_ind)
                if ind_key not in created_flags['indicadores']:
                    created_flags['indicadores'].add(ind_key)
                    stats['indicadores'] += 1

    return stats


PLANTILLA_CSV = (
    'id_eje,eje,id_plan,plan,proposito_plan,vision_plan,id_programa,programa,objetivo,proyecto,indicador,unidad_medida,frecuencia\r\n'
    '1,Fortalecimiento de la Economía del Conocimiento,1,1. Fortalecimiento...,Posicionar a la provincia...,Consolidar hub...,1.1,Impulso al Emprendedurismo Tech,Acelerar startups.,Proyecto piloto tech,Indicador de startups creadas,Unidades,Anual\r\n'
)
